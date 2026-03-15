"""
indiapost_history_pdf.py  —  Generate one PDF per article with full event history.

Usage:
  python indiapost_history_pdf.py input.csv
  python indiapost_history_pdf.py input.xlsx
  python indiapost_history_pdf.py              (prompts for file path)

Output folder: History_PDFs/  (created next to the input file)
Each PDF is named  <ArticleNumber>.pdf  and contains:
  - Booking summary (article type, booked at/on, origin → destination)
  - Delivery status
  - Full chronological event table (Date | Time | Event | Office)
"""

import sys
import os
import csv
import json
import openpyxl
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    from fpdf import FPDF
except ImportError:
    sys.exit("[ERROR] fpdf2 not installed. Run: pip install fpdf2")

from indiapost_api import get_receipt_json, _auth

# ── Layout constants ──────────────────────────────────────────────────────────
MARGIN      = 15
PAGE_W      = 210   # A4
CONTENT_W   = PAGE_W - 2 * MARGIN

CLR_NAVY    = (31,  78, 121)
CLR_GREEN   = (84, 130,  53)
CLR_RED     = (192,  0,  0)
CLR_ORANGE  = (197, 90,  17)
CLR_BLUE    = ( 31, 73, 125)
CLR_GREY    = (166,166,166)
CLR_HDRROW  = (217,226,243)
CLR_ALTROW  = (242,242,242)


def status_colour(del_status: str, events: list) -> tuple:
    ds = del_status.lower()
    if ds == "delivered":
        top = events[0].get("event", "") if events else ""
        if "Addressee" in top:   return CLR_GREEN
        return CLR_BLUE
    # check if return journey
    LOGISTICS = {"Bag Close","Bag Dispatch","Bag Received","Item Invoiced","Item Book","Item Received"}
    for e in events:
        ev = e.get("event","")
        if ev not in LOGISTICS:
            if "Return" in ev: return CLR_ORANGE
            break
    return CLR_GREY


def delivery_label(del_status: str, events: list) -> str:
    ds = del_status.lower()
    if ds == "delivered":
        top = events[0].get("event", "") if events else ""
        if "Addressee" in top: return "DELIVERED TO ADDRESSEE"
        return "DELIVERED"
    LOGISTICS = {"Bag Close","Bag Dispatch","Bag Received","Item Invoiced","Item Book","Item Received"}
    for e in events:
        ev = e.get("event","")
        if ev not in LOGISTICS:
            if "Return" in ev: return "RETURN JOURNEY"
            if "Onhold" in ev: return "ON HOLD"
            break
    return "IN TRANSIT"


class HistoryPDF(FPDF):
    def __init__(self, article_number: str):
        super().__init__()
        self.article_number = article_number
        self.set_margins(MARGIN, MARGIN, MARGIN)
        self.set_auto_page_break(auto=True, margin=MARGIN)

    def header(self):
        # Top bar
        self.set_fill_color(*CLR_NAVY)
        self.rect(0, 0, PAGE_W, 14, "F")
        self.set_text_color(255, 255, 255)
        self.set_font("Helvetica", "B", 11)
        self.set_xy(MARGIN, 3)
        self.cell(0, 8, "India Post  -  Article Tracking History")
        self.set_font("Helvetica", "", 8)
        self.set_xy(MARGIN, 3)
        self.cell(CONTENT_W, 8, f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}",
                  align="R")
        self.set_text_color(0, 0, 0)
        self.ln(14)

    def footer(self):
        self.set_y(-12)
        self.set_font("Helvetica", "", 7)
        self.set_text_color(150, 150, 150)
        self.cell(0, 5, f"Article: {self.article_number}   |   Page {self.page_no()}", align="C")
        self.set_text_color(0, 0, 0)


def build_pdf(receipt: dict, out_path: str):
    num     = receipt["article_number"]
    events  = receipt["events"]          # newest first from API
    chron   = list(reversed(events))     # chronological for display

    pdf = HistoryPDF(num)
    pdf.add_page()
    pdf.set_font("Helvetica", "", 10)

    # ── Article number banner ─────────────────────────────────────────────────
    pdf.set_fill_color(*CLR_NAVY)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 14)
    pdf.cell(CONTENT_W, 10, f"  {num}", fill=True, ln=True)
    pdf.ln(4)

    # ── Booking summary box ───────────────────────────────────────────────────
    pdf.set_text_color(0, 0, 0)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(*CLR_HDRROW)
    pdf.cell(CONTENT_W, 6, "  BOOKING DETAILS", fill=True, ln=True)
    pdf.set_font("Helvetica", "", 9)

    col_w = CONTENT_W / 2
    rows = [
        ("Article Type",        receipt["article_type"]),
        ("Booked At",           receipt["booked_at"]),
        ("Booked On",           receipt["booked_on"]),
        ("Origin PIN",          receipt["origin_pincode"]),
        ("Destination PIN",     receipt["destination_pincode"]),
        ("Delivery Location",   receipt["delivery_location"]),
        ("Delivery Confirmed",  receipt["delivery_confirmed_on"] or "-"),
        ("Tariff (Rs.)",        str(receipt["tariff"]) if receipt["tariff"] else "-"),
    ]
    for i, (label, value) in enumerate(rows):
        fill_clr = CLR_ALTROW if i % 2 == 0 else (255, 255, 255)
        pdf.set_fill_color(*fill_clr)
        pdf.set_font("Helvetica", "B", 8)
        pdf.cell(col_w * 0.45, 6, f"  {label}", fill=True, border=0)
        pdf.set_font("Helvetica", "", 8)
        pdf.cell(CONTENT_W - col_w * 0.45, 6, str(value or "-"), fill=True, border=0, ln=True)

    pdf.ln(4)

    # ── Delivery status pill ──────────────────────────────────────────────────
    label   = delivery_label(receipt["del_status"], events)
    clr     = status_colour(receipt["del_status"], events)
    pdf.set_fill_color(*clr)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(CONTENT_W, 8, f"  STATUS:  {label}", fill=True, ln=True)
    pdf.set_text_color(0, 0, 0)
    pdf.ln(4)

    # ── Event history table ───────────────────────────────────────────────────
    pdf.set_fill_color(*CLR_HDRROW)
    pdf.set_font("Helvetica", "B", 9)
    pdf.cell(CONTENT_W, 6, f"  EVENT HISTORY  ({len(chron)} events, oldest first)", fill=True, ln=True)

    # Column widths
    cw = [28, 22, 65, CONTENT_W - 28 - 22 - 65]   # Date | Time | Event | Office

    # Table header
    pdf.set_fill_color(*CLR_NAVY)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 8)
    for txt, w in zip(["Date", "Time", "Event", "Office"], cw):
        pdf.cell(w, 6, f"  {txt}", fill=True, border=0)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)

    for idx, ev in enumerate(chron):
        fill_clr = CLR_ALTROW if idx % 2 == 0 else (255, 255, 255)
        pdf.set_fill_color(*fill_clr)
        pdf.set_font("Helvetica", "", 8)

        date_str  = ev.get("date", "")
        time_str  = ev.get("time", "")
        evt_name  = ev.get("event", "")
        office    = ev.get("office", "")

        # Bold delivery/return events
        if "Delivered" in evt_name or "Return" in evt_name:
            pdf.set_font("Helvetica", "B", 8)
            pdf.set_text_color(*CLR_GREEN if "Delivered" in evt_name else CLR_ORANGE)

        for txt, w in zip([date_str, time_str, evt_name, office], cw):
            pdf.cell(w, 5.5, f"  {txt}", fill=True, border=0)
        pdf.ln()
        pdf.set_text_color(0, 0, 0)
        pdf.set_font("Helvetica", "", 8)

    pdf.output(out_path)


# ── Input loading (shared with indiapost_report) ──────────────────────────────

def load_ids(path: str) -> list[str]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            col = _find_col(headers)
            return [str(r.get(col, "") or "").strip() for r in reader
                    if str(r.get(col, "") or "").strip()]
    elif ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        col = _find_col(headers)
        col_idx = headers.index(col)
        ids = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            v = str(row[col_idx] or "").strip()
            if v:
                ids.append(v)
        return ids
    else:
        sys.exit(f"[ERROR] Unsupported file: {ext}")


def _find_col(cols: list) -> str:
    candidates = ["tracking id","tracking_id","article no","article number",
                  "articleno","trackingnumber","consignment no",
                  "article numbers","article_numbers","consignment number"]
    for c in cols:
        if c and str(c).strip().lower() in candidates:
            return c
    return cols[0] if cols else ""


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) > 1:
        input_path = sys.argv[1]
    else:
        input_path = input("Enter path to input CSV/Excel file:\n> ").strip().strip('"')

    if not os.path.exists(input_path):
        sys.exit(f"[ERROR] File not found: {input_path}")

    ids = load_ids(input_path)
    if not ids:
        sys.exit("[ERROR] No tracking IDs found.")

    print(f"\n[1/3] Loaded {len(ids)} article IDs from {os.path.basename(input_path)}")

    out_dir = os.path.join(os.path.dirname(os.path.abspath(input_path)), "History_PDFs")
    os.makedirs(out_dir, exist_ok=True)

    print(f"[2/3] Fetching tracking data...")
    auth     = _auth()
    receipts = get_receipt_json(ids, auth=auth)
    print(f"      Got data for {len(receipts)}/{len(ids)} articles.")

    print(f"[3/3] Generating PDFs -> {out_dir}")
    ok = 0
    for id_ in ids:
        receipt = receipts.get(id_)
        if not receipt:
            print(f"  [SKIP] {id_} — no data")
            continue
        out_path = os.path.join(out_dir, f"{id_}.pdf")
        try:
            build_pdf(receipt, out_path)
            ok += 1
        except Exception as e:
            print(f"  [ERROR] {id_}: {e}")

    print(f"\n[DONE] {ok}/{len(ids)} PDFs saved to: {out_dir}")

    # Also save a combined receipts.json for reference
    json_path = os.path.join(out_dir, "receipts.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(receipts, f, indent=2, ensure_ascii=False)
    print(f"       receipts.json also saved -> {json_path}")


if __name__ == "__main__":
    main()
