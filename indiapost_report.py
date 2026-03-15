"""
IndiaPost Tracking Report Generator
Reads tracking IDs from a CSV or Excel file, fetches status from the
IndiaPost Production API, and writes an Excel report.

Input file format (CSV or .xlsx):
  - Must have a column named: Tracking ID  (or tracking_id / Article No)
  - Optional columns carried through to output: Name, Case No, etc.

Usage:
  python indiapost_report.py input.csv
  python indiapost_report.py input.xlsx
  python indiapost_report.py              (prompts for file path)

Output: IndiaPost_Report_<timestamp>.xlsx in the same folder as the input.
"""

import sys
import os
import math
import csv
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("[ERROR] openpyxl not installed. Run: pip install openpyxl")

from indiapost_tracker import AuthManager, fetch_tracking, determine_delivery_outcome, CLIENT_ID, CLIENT_SECRET

BATCH_SIZE = 20   # Smaller batches — API drops tracking_details on larger batches

# ── Colour palette ────────────────────────────────────────────────────────────
CLR_HEADER      = "1F4E79"   # dark blue
CLR_ADDR        = "C6EFCE"   # light green      — Delivered to Addressee
CLR_OFFICE      = "E2EFDA"   # pale green       — Delivered at Office
CLR_RETURN_DONE = "FFC7CE"   # light red        — Returned to Sender (final)
CLR_RETURN_JOUR = "FFCCBB"   # light orange-red — Return Journey (in progress)
CLR_ONWARD      = "DDEBF7"   # light blue       — Onward Journey
CLR_ONHOLD      = "FFF2CC"   # pale yellow      — On Hold
CLR_NODATA      = "F2F2F2"   # grey             — No data / No events


def load_input(path: str) -> tuple[list[dict], list[str]]:
    """Return (rows, all_column_names) from a CSV or Excel input file."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows = list(reader)
            cols = reader.fieldnames or []
    elif ext in (".xlsx", ".xlsm"):
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = wb.active
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        rows, cols = [], headers
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(v is not None for v in row):
                rows.append(dict(zip(headers, row)))
    else:
        sys.exit(f"[ERROR] Unsupported file type: {ext}  (use .csv or .xlsx)")
    return rows, list(cols)


def find_tracking_col(cols: list[str]) -> str:
    """Identify which column holds the tracking ID."""
    candidates = ["tracking id", "tracking_id", "article no", "article number",
                  "articleno", "trackingnumber", "consignment no", "article numbers",
                  "article_numbers", "consignment number"]
    for col in cols:
        if col and col.strip().lower() in candidates:
            return col
    # fallback: first column
    return cols[0] if cols else ""


def fetch_all(auth: AuthManager, ids: list[str]) -> dict[str, dict]:
    """
    Fetch tracking for all IDs in batches; return dict keyed by article_number.
    The production API sometimes returns tracking_details=null for some articles
    in larger batches. Those are retried individually to get the real data.
    """
    result_map = {}
    batches = math.ceil(len(ids) / BATCH_SIZE)

    import time as _time
    for i in range(batches):
        batch = ids[i * BATCH_SIZE : (i + 1) * BATCH_SIZE]
        print(f"  Fetching batch {i+1}/{batches}  ({len(batch)} articles)...", end=" ", flush=True)
        for attempt in range(3):
            try:
                resp = fetch_tracking(auth, batch)
                break
            except Exception as e:
                if attempt < 2:
                    print(f"[timeout, retrying...]", end=" ", flush=True)
                    _time.sleep(5)
                else:
                    print(f"[failed after 3 attempts: {e}]")
                    resp = {}
        for art in (resp.get("data") or []):
            num = (art.get("booking_details") or {}).get("article_number", "")
            if num:
                result_map[num] = art
        print("done")

    # Retry individually any that came back with null tracking_details
    missing_events = [
        id_ for id_ in ids
        if id_ in result_map and not result_map[id_].get("tracking_details")
    ]
    if missing_events:
        print(f"  Retrying {len(missing_events)} articles with null tracking data...")
        for id_ in missing_events:
            resp = fetch_tracking(auth, [id_])
            for art in (resp.get("data") or []):
                num = (art.get("booking_details") or {}).get("article_number", "")
                if num and art.get("tracking_details"):
                    result_map[num] = art

    return result_map


def row_colour(outcome: str) -> str:
    if "Addressee"       in outcome: return CLR_ADDR
    if "at Office"       in outcome: return CLR_OFFICE
    if "Returned to"     in outcome: return CLR_RETURN_DONE
    if "Return Journey"  in outcome: return CLR_RETURN_JOUR
    if "Onward Journey"  in outcome: return CLR_ONWARD
    if "On Hold"         in outcome: return CLR_ONHOLD
    return CLR_NODATA


def write_report(input_rows: list[dict], extra_cols: list[str],
                 tracking_col: str, data_map: dict[str, dict],
                 out_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tracking Report"

    # ── Output columns ────────────────────────────────────────────────────────
    fixed_out = [
        "Article No",
        "Booked At",
        "Booked On",
        "Origin PIN",
        "Destination PIN",
        "Delivery Location",
        "Last Event",
        "Last Event Date",
        "Last Event Office",
        "Delivery Outcome",
    ]
    pass_cols = [c for c in extra_cols if c and c != tracking_col]
    all_cols  = [tracking_col] + pass_cols + fixed_out

    # ── Header row ────────────────────────────────────────────────────────────
    hdr_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    hdr_fill  = PatternFill("solid", fgColor=CLR_HEADER)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.append(all_cols)
    for cell in ws[1]:
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align
        cell.border    = border
    ws.row_dimensions[1].height = 30

    # ── Data rows ─────────────────────────────────────────────────────────────
    for input_row in input_rows:
        tracking_id = str(input_row.get(tracking_col) or "").strip()
        art         = data_map.get(tracking_id, {})
        booking     = art.get("booking_details") or {}
        events      = art.get("tracking_details") or []
        latest      = events[0] if events else {}
        outcome     = determine_delivery_outcome(art) if art else "Not Found"

        booked_on_raw = booking.get("booked_on", "") or ""
        booked_on     = booked_on_raw[:10] if booked_on_raw else ""

        last_date_raw = latest.get("date", "") or ""
        last_date     = last_date_raw[:10] if last_date_raw else ""

        row_data = [tracking_id] + [input_row.get(c, "") for c in pass_cols] + [
            booking.get("article_number", tracking_id),
            booking.get("booked_at", ""),
            booked_on,
            booking.get("origin_pincode", ""),
            booking.get("destination_pincode", ""),
            booking.get("delivery_location", ""),
            latest.get("event", ""),
            last_date,
            latest.get("office", ""),
            outcome,
        ]
        ws.append(row_data)

        # colour the entire row by outcome
        fill = PatternFill("solid", fgColor=row_colour(outcome))
        row_num = ws.max_row
        for cell in ws[row_num]:
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(vertical="center")
            cell.font      = Font(name="Calibri", size=10)

        # bold the outcome cell
        ws.cell(row=row_num, column=len(all_cols)).font = Font(
            name="Calibri", size=10, bold=True
        )

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        tracking_col:         22,
        "Article No":         22,
        "Booked At":          22,
        "Booked On":          14,
        "Origin PIN":         13,
        "Destination PIN":    16,
        "Delivery Location":  25,
        "Last Event":         28,
        "Last Event Date":    16,
        "Last Event Office":  25,
        "Delivery Outcome":   30,
    }
    for i, col in enumerate(all_cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = col_widths.get(col, 18)

    # ── Freeze header + auto-filter ───────────────────────────────────────────
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Legend sheet ─────────────────────────────────────────────────────────
    legend = wb.create_sheet("Legend")
    items = [
        (CLR_ADDR,        "Delivered to Addressee",  "Successfully delivered to the intended recipient."),
        (CLR_OFFICE,      "Delivered at Office",     "Delivered but not directly to addressee (e.g. held at post office)."),
        (CLR_RETURN_DONE, "Returned to Sender",      "Item fully returned — delivery failed."),
        (CLR_RETURN_JOUR, "Return Journey",           "Delivery attempted but failed; item is now heading back."),
        (CLR_ONWARD,      "Onward Journey",           "Item is in transit towards the destination."),
        (CLR_ONHOLD,      "On Hold",                  "Item is on hold at a facility."),
        (CLR_NODATA,      "Not Found / No Events",    "No tracking data returned by the API."),
    ]
    legend.append(["Colour", "Status", "Meaning"])
    for cell in legend[1]:
        cell.font   = Font(bold=True, name="Calibri")
        cell.border = border
    for clr, status, meaning in items:
        legend.append(["", status, meaning])
        r = legend.max_row
        legend.cell(r, 1).fill = PatternFill("solid", fgColor=clr)
        for c in range(1, 4):
            legend.cell(r, c).border = border
            legend.cell(r, c).font   = Font(name="Calibri", size=10)
    legend.column_dimensions["A"].width = 10
    legend.column_dimensions["B"].width = 28
    legend.column_dimensions["C"].width = 55

    wb.save(out_path)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    if len(sys.argv) > 1:
        input_path = sys.argv[1]
    else:
        input_path = input("Enter path to input CSV/Excel file:\n> ").strip().strip('"')

    if not os.path.exists(input_path):
        sys.exit(f"[ERROR] File not found: {input_path}")

    print(f"\n[1/4] Loading input: {input_path}")
    rows, cols = load_input(input_path)
    tracking_col = find_tracking_col(cols)
    print(f"      Tracking ID column : '{tracking_col}'")
    print(f"      Rows loaded         : {len(rows)}")

    ids = [str(r.get(tracking_col) or "").strip() for r in rows if r.get(tracking_col)]
    ids = [i for i in ids if i]
    if not ids:
        sys.exit("[ERROR] No tracking IDs found in the input file.")

    print(f"\n[2/4] Authenticating...")
    auth = AuthManager(CLIENT_ID, CLIENT_SECRET)
    auth.get_token()

    print(f"\n[3/4] Fetching tracking data for {len(ids)} article(s)...")
    data_map = fetch_all(auth, ids)
    found = len(data_map)
    print(f"      Data returned for {found}/{len(ids)} articles.")

    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir    = os.path.dirname(os.path.abspath(input_path))
    out_path   = os.path.join(out_dir, f"IndiaPost_Report_{timestamp}.xlsx")

    print(f"\n[4/4] Writing report: {out_path}")
    write_report(rows, cols, tracking_col, data_map, out_path)
    print(f"\n[DONE] Report saved -> {out_path}")

    # Print a quick summary
    outcomes = [determine_delivery_outcome(data_map[i]) if i in data_map else "Not Found"
                for i in ids]
    counts = {
        "Delivered to Addressee" : sum(1 for o in outcomes if "Addressee"      in o),
        "Delivered at Office"    : sum(1 for o in outcomes if "at Office"       in o),
        "Returned to Sender"     : sum(1 for o in outcomes if "Returned to"     in o),
        "Return Journey"         : sum(1 for o in outcomes if "Return Journey"  in o),
        "Onward Journey"         : sum(1 for o in outcomes if "Onward Journey"  in o),
        "On Hold"                : sum(1 for o in outcomes if "On Hold"         in o),
        "Not Found / No Events"  : sum(1 for o in outcomes if o in ("Not Found", "No Events")),
    }
    print(f"\n  Summary ({len(ids)} articles):")
    for label, count in counts.items():
        if count:
            print(f"    {label:<28}: {count}")


if __name__ == "__main__":
    main()
