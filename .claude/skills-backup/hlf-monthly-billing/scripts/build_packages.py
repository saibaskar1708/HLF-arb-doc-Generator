#!/usr/bin/env python3
"""
HLF Monthly Billing Package Builder
=====================================
Generates per-state invoice packages for Hinduja Leyland Finance Ltd.

Each package = Covering Letter PDF + GST Bill PDF + Excel Case Details PDF (merged)

Usage:
    python build_packages.py \
        --billing-excel  <path-to-billing-data.xlsx> \
        --gst-bills-dir  <path-to-folder-with-state-GST-pdfs> \
        --output-dir     <where-to-save-packages> \
        --bill-date      14.03.2026 \
        --billing-month  MAR \
        --billing-year   2026

Optional:
    --states-only        Comma-separated list of states to process (default: all)
    --skip-missing-gst   Skip states with no GST bill (default: auto-generate from data)
"""

import argparse, json, os, shutil, subprocess, sys
from pathlib import Path
import pandas as pd
from weasyprint import HTML
from pypdf import PdfWriter, PdfReader
from num2words import num2words
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Helpers ──────────────────────────────────────────────────────────────────

SCRIPT_DIR = Path(__file__).parent
MASTER = json.loads((SCRIPT_DIR / "state_master.json").read_text())
CO = MASTER["our_company"]
STATES = MASTER["states"]

MONTH_ABBR = {
    "JAN": "January", "FEB": "February", "MAR": "March", "APR": "April",
    "MAY": "May", "JUN": "June", "JUL": "July", "AUG": "August",
    "SEP": "September", "OCT": "October", "NOV": "November", "DEC": "December"
}

def amount_in_words(amount):
    try:
        rupees = int(amount)
        paise_val = round((amount - rupees) * 100)
        words = num2words(rupees, lang="en_IN").title().replace(",", "")
        if paise_val > 0:
            pw = num2words(paise_val, lang="en_IN").title().replace(",", "")
            return f"INR {words} Rupees And {pw} Paise Only"
        return f"INR {words} Rupees Only"
    except Exception:
        return f"INR {amount:.2f}"

def get_invoice_no(state, billing_month, billing_year):
    sd = STATES[state]
    yr2 = str(billing_year)[-2:]
    return f"{sd['inv_prefix']}/{billing_month}/{sd['state_suffix']}"

def get_groups(df, state):
    """Return case groupings for a state: list of dicts with desc, sub, count, fee"""
    sdf = df[df["State"] == state].copy()
    groups = (
        sdf.groupby(["Type of Case", "Stage of Case"])
        .agg(Count=("Contract No", "count"), Fee=("Fees Amount", "first"))
        .reset_index()
    )
    result = []
    for _, row in groups.iterrows():
        result.append({
            "desc": "Professional / Legal & Arbitration Services",
            "sub": f"Total {int(row['Count'])} '{row['Type of Case']}' Type Cases, {row['Stage of Case']}",
            "count": int(row["Count"]),
            "fee": float(row["Fee"]),
        })
    return result, sdf

# ── Covering Letter HTML ──────────────────────────────────────────────────────

COMMON_CSS = """
@page { size: A4; margin: 15mm 15mm 15mm 15mm; }
body { font-family: Arial, sans-serif; font-size: 10pt; color: #000; margin: 0; }
table { border-collapse: collapse; width: 100%; }
td, th { padding: 4px 6px; }
.bold { font-weight: bold; }
.center { text-align: center; }
.right { text-align: right; }
"""

def build_covering_letter_html(state, df, bill_date, billing_month, billing_year):
    sd = STATES[state]
    groups, sdf = get_groups(df, state)
    total_fees = sdf["Fees Amount"].sum()
    inv_no = get_invoice_no(state, billing_month, billing_year)
    month_label = MONTH_ABBR.get(billing_month, billing_month)

    if sd["gst_type"] == "CGST+SGST":
        cgst = round(total_fees * 0.09, 2)
        sgst = round(total_fees * 0.09, 2)
        gst_amount = cgst + sgst
        gst_label = "GST (CGST @9% &amp; SGST @9%)"
    else:
        gst_amount = round(total_fees * 0.18, 2)
        gst_label = "GST (IGST @18%)"

    grand_total = total_fees + gst_amount
    addr_lines = "<br>".join(sd["hlf_address"])

    rows_html = ""
    for i, g in enumerate(groups, 1):
        subtotal = g["count"] * g["fee"]
        rows_html += f"""
        <tr style="background:{'#EBF3FB' if i%2==0 else '#fff'}">
          <td class="center">{i}.</td>
          <td>{g['desc']}<br><span style="font-size:8.5pt;color:#333">{g['sub']}</span></td>
          <td class="center">{g['count']}</td>
          <td class="right">&#8377; {g['fee']:,.0f}</td>
          <td class="right">&#8377; {subtotal:,.2f}</td>
        </tr>"""

    gst_row = f"""
        <tr style="background:#FFF2CC">
          <td colspan="4" class="right bold">{gst_label}</td>
          <td class="right bold">&#8377; {gst_amount:,.2f}</td>
        </tr>"""

    total_row = f"""
        <tr style="background:#1F4E79;color:white">
          <td colspan="4" class="right bold">TOTAL</td>
          <td class="right bold">&#8377; {grand_total:,.2f}</td>
        </tr>"""

    words = amount_in_words(grand_total)

    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<style>
{COMMON_CSS}
.letterhead {{ border-bottom: 3px solid #1F4E79; padding-bottom: 8px; margin-bottom: 12px; }}
.lh-title {{ font-size: 16pt; font-weight: bold; color: #1F4E79; }}
.lh-sub {{ font-size: 9pt; color: #555; }}
.ref-table td {{ border: none; padding: 2px 4px; }}
.main-table th {{ background: #1F4E79; color: white; font-size: 9pt; text-align: center; border: 1px solid #ccc; }}
.main-table td {{ border: 1px solid #ccc; font-size: 9pt; vertical-align: top; }}
.bank-table td {{ border: 1px solid #999; padding: 6px 10px; vertical-align: top; font-size: 9pt; }}
.words-box {{ background: #EBF3FB; border: 1px solid #1F4E79; padding: 6px 10px; font-size: 9pt; margin: 8px 0; font-style: italic; }}
.sig-space {{ height: 60px; }}
</style></head><body>

<div class="letterhead">
  <table><tr>
    <td><div class="lh-title">{CO['name']}</div>
        <div class="lh-sub">{CO['address']}</div>
        <div class="lh-sub">GSTIN/UIN: {CO['gstin']} &nbsp;|&nbsp; Ph: {CO['phone']} &nbsp;|&nbsp; Email: {CO['email']}</div>
    </td>
  </tr></table>
</div>

<table class="ref-table" style="margin-bottom:10px">
  <tr>
    <td><b>Bill No:</b> {inv_no}</td>
    <td class="right"><b>Date:</b> {bill_date}</td>
  </tr>
</table>

<div style="margin-bottom:10px; line-height:1.6">
<b>To,</b><br>
{addr_lines}<br>
<b>GSTIN/UIN:</b> {sd['gstin']}<br>
<b>State Name:</b> {state}, Code: {sd['state_code']}
</div>

<div style="margin-bottom:6px"><b>Kind Attention:</b> {MASTER['hlf_contact']}</div>

<div style="margin-bottom:10px">
Request you to kindly process the Arbitration Charges for the Arbitration Cases
initiated during <b>{month_label} {billing_year}</b> as per the annexure.
</div>

<div style="margin-bottom:8px"><b>Sub: Arbitration Charges &ndash; {state} ({month_label} {billing_year})</b></div>

<table class="main-table">
  <tr>
    <th style="width:5%">S.No.</th>
    <th style="width:45%">Description</th>
    <th style="width:12%">Total Count</th>
    <th style="width:17%">Arbitration Fees per Case</th>
    <th style="width:21%">Total</th>
  </tr>
  {rows_html}
  {gst_row}
  {total_row}
</table>

<div class="words-box">{words}</div>

<div style="margin-bottom:12px; font-size:9.5pt">
Kindly release the above Arbitration Fees amount at the earliest.
</div>

<table class="bank-table">
  <tr>
    <td style="width:60%">
      <b>Bank Details &ndash;</b><br><br>
      A/C Holder&rsquo;s Name: <b>{CO['name']}</b><br>
      Bank Name: {CO['bank_name']}<br>
      A/C No.: {CO['bank_ac']}<br>
      Branch &amp; IFS Code: {CO['bank_branch']} &amp; {CO['bank_ifsc']}<br>
      GSTIN/UIN: {CO['gstin']}
    </td>
    <td style="width:40%; text-align:center; vertical-align:bottom">
      <b>Yours Faithfully,</b><br>
      <div class="sig-space"></div>
      <b>{CO['signatory']}</b><br>
      <span style="font-size:8.5pt">{CO['name']}</span>
    </td>
  </tr>
</table>

</body></html>"""

# ── Generate GST Bill for a state (fallback when PDF not provided) ────────────

def build_gst_bill_html(state, df, bill_date, billing_month, billing_year):
    sd = STATES[state]
    groups, sdf = get_groups(df, state)
    total_fees = sdf["Fees Amount"].sum()
    inv_no = get_invoice_no(state, billing_month, billing_year)

    if sd["gst_type"] == "CGST+SGST":
        gst_amount = round(total_fees * 0.18, 2)
        tax_label = "CGST @ 9% + SGST @ 9%"
        tax_col_header = "CGST + SGST @ 18%"
    else:
        gst_amount = round(total_fees * 0.18, 2)
        tax_label = "Output IGST @ 18%"
        tax_col_header = "IGST TAX @ 18%"

    grand_total = total_fees + gst_amount
    words = amount_in_words(grand_total)
    tax_words = amount_in_words(gst_amount)

    # Convert bill_date from DD.MM.YYYY to DD-Mon-YYYY
    try:
        from datetime import datetime
        dt = datetime.strptime(bill_date, "%d.%m.%Y")
        formatted_date = dt.strftime("%d-%b-%Y")
    except Exception:
        formatted_date = bill_date

    rows_html = ""
    for i, g in enumerate(groups, 1):
        subtotal = g["count"] * g["fee"]
        rows_html += f"""
        <tr>
          <td class="center">{i}</td>
          <td>{g['desc']}<br><i style="font-size:8pt">{g['sub']}</i></td>
          <td class="center">{g['count']}</td>
          <td class="right">{g['fee']:,.0f}</td>
          <td class="center">998215</td>
          <td class="right">{subtotal:,.2f}</td>
        </tr>"""

    addr_lines_buyer = "<br>".join(sd["hlf_address"][1:])  # skip the first line (M/S...)

    return f"""<!DOCTYPE html><html><head><meta charset="UTF-8">
<style>
@page {{ size: A4; margin: 12mm; }}
body {{ font-family: Arial, sans-serif; font-size: 9pt; color: #000; margin: 0; }}
table {{ border-collapse: collapse; width: 100%; }}
.outer td {{ border: 1px solid #000; padding: 4px 6px; vertical-align:top; }}
.outer th {{ border: 1px solid #000; padding: 4px 6px; background: #e8e8e8; font-size: 8.5pt; }}
.title {{ font-size: 14pt; font-weight: bold; text-align: center; padding: 6px 0 2px; }}
.sub-title {{ font-size: 10pt; text-align: center; padding: 0 0 6px; }}
.right {{ text-align: right; }} .center {{ text-align: center; }} .bold {{ font-weight: bold; }}
.shade {{ background: #f0f0f0; }}
</style></head><body>
<div class="title">Tax Invoice</div>
<div class="sub-title bold">{CO['name']}</div>
<table class="outer" style="margin-top:4px">
  <tr>
    <td colspan="3" style="width:60%">
      {CO['address']}<br>
      <b>GSTIN/UIN:</b> {CO['gstin']} &nbsp;|&nbsp; State: {CO['state']}, Code: {CO['state_code']}<br>
      Contact: {CO['phone']} &nbsp;|&nbsp; E-Mail: {CO['email']}
    </td>
    <td colspan="3" style="width:40%">
      <b>Invoice No.:</b> {inv_no}<br><br>
      <b>Dated:</b> {formatted_date}
    </td>
  </tr>
  <tr>
    <td colspan="4">
      <b>Buyer (Bill to):</b><br>
      M/S. HINDUJA LEYLAND FINANCE LIMITED<br>(PAN: {MASTER['hlf_pan']})<br>
      {addr_lines_buyer}<br>
      <b>GSTIN/UIN:</b> {sd['gstin']}<br>
      <b>State Name:</b> {state}, Code: {sd['state_code']}
    </td>
    <td colspan="2"><b>Court Name:</b> {MASTER['court_name']}</td>
  </tr>
  <tr class="shade">
    <th>Sl No</th><th>Particulars</th><th>Total Count</th>
    <th>Fees</th><th>HSN/SAC</th><th>Amount</th>
  </tr>
  {rows_html}
  <tr class="shade">
    <td colspan="5" class="right bold">{tax_label}</td>
    <td class="right bold">{gst_amount:,.2f}</td>
  </tr>
  <tr>
    <td colspan="4" style="font-size:8pt">Whether Tax is payable on Reverse Charge: No</td>
    <td class="right bold">Total</td>
    <td class="right bold">{grand_total:,.2f}</td>
  </tr>
  <tr>
    <td colspan="4"><b>Amount Chargeable (in words)</b><br><i>{words}</i></td>
    <td colspan="2" class="center">E. &amp; O.E</td>
  </tr>
  <tr class="shade">
    <th colspan="2">HSN/SAC</th><th>Taxable Value</th>
    <th>Rate</th><th colspan="2">{tax_col_header} &nbsp; Amount &nbsp; Total Tax</th>
  </tr>
  <tr>
    <td colspan="2" class="center">998215</td>
    <td class="right">{total_fees:,.2f}</td>
    <td class="center">18%</td>
    <td colspan="2" class="right">{gst_amount:,.2f} &nbsp;&nbsp; {gst_amount:,.2f}</td>
  </tr>
  <tr class="shade">
    <td colspan="2" class="bold">Total</td>
    <td class="right bold">{total_fees:,.2f}</td>
    <td></td>
    <td colspan="2" class="right bold">{gst_amount:,.2f} &nbsp;&nbsp; {gst_amount:,.2f}</td>
  </tr>
  <tr>
    <td colspan="6"><b>Tax Amount (in words):</b> <i>{tax_words}</i></td>
  </tr>
  <tr>
    <td colspan="3">
      <b>Company's Bank Details</b><br>
      A/c Holder's Name: {CO['name']}<br>
      Bank Name: {CO['bank_name']}<br>
      A/c No.: {CO['bank_ac']}<br>
      Branch &amp; IFS Code: {CO['bank_branch']} &amp; {CO['bank_ifsc']}
    </td>
    <td colspan="3" class="center">
      <b>For {CO['name']}</b><br><br><br><br>
      <b>Authorised Signatory</b>
    </td>
  </tr>
  <tr><td colspan="6" class="center" style="font-size:8pt">This is a Computer Generated Invoice</td></tr>
</table>
</body></html>"""

# ── Excel Case Detail Sheet ───────────────────────────────────────────────────

def build_excel_case_detail(state, df, temp_dir, billing_month, billing_year):
    sdf = df[df["State"] == state].copy().reset_index(drop=True)
    sdf["S No"] = range(1, len(sdf) + 1)
    cols = ["S No", "Contract No", "Customer Name", "Case No",
            "Type of Case", "Stage of Case", "Court Name", "Fees Amount"]
    sdf = sdf[cols]

    wb = Workbook()
    ws = wb.active
    ws.title = state[:31]
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True

    month_label = MONTH_ABBR.get(billing_month, billing_month)
    ws.merge_cells("A1:H1")
    tc = ws["A1"]
    tc.value = f"Arbitration Case Details \u2013 {state} | Hinduja Leyland Finance Ltd. | {month_label} {billing_year}"
    tc.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    tc.fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    tc.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    thin = Side(style="thin", color="AAAAAA")
    bdr = Border(left=thin, right=thin, top=thin, bottom=thin)
    headers = list(sdf.columns)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font = Font(name="Arial", bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", start_color="2E75B6", end_color="2E75B6")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = bdr
    ws.row_dimensions[2].height = 18

    for ri, row in sdf.iterrows():
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri + 3, column=ci, value=val)
            c.font = Font(name="Arial", size=9)
            bg = "EBF3FB" if ri % 2 == 0 else "FFFFFF"
            c.fill = PatternFill("solid", start_color=bg, end_color=bg)
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = bdr
            if headers[ci - 1] == "Fees Amount":
                c.number_format = "#,##0"

    last_dr = 2 + len(sdf)
    fees_col = headers.index("Fees Amount") + 1
    total_fees = sdf["Fees Amount"].sum()
    gst_amount = round(total_fees * 0.18, 2)
    grand_total = total_fees + gst_amount

    for i, (lbl, val) in enumerate([
        ("Total Fees", total_fees), ("GST @ 18%", gst_amount), ("Grand Total", grand_total)
    ]):
        r = last_dr + 2 + i
        lc = ws.cell(row=r, column=fees_col - 1, value=lbl)
        lc.font = Font(name="Arial", bold=True, size=9)
        lc.alignment = Alignment(horizontal="right")
        lc.border = bdr
        vc = ws.cell(row=r, column=fees_col, value=val)
        vc.font = Font(name="Arial", bold=True, size=9,
                       color="1F4E79" if lbl == "Grand Total" else "000000")
        vc.number_format = "#,##0.00"
        vc.border = bdr
        vc.alignment = Alignment(horizontal="right")

    widths = {"S No": 5, "Contract No": 16, "Customer Name": 22, "Case No": 20,
              "Type of Case": 10, "Stage of Case": 22, "Court Name": 22, "Fees Amount": 13}
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(h, 14)

    ws.freeze_panes = "A3"
    safe = state.replace(" ", "_")
    path = os.path.join(temp_dir, f"xl_{safe}.xlsx")
    wb.save(path)
    return path

def excel_to_pdf(xlsx_path, out_pdf, temp_dir):
    """Convert Excel to PDF using LibreOffice"""
    r = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", temp_dir, xlsx_path],
        capture_output=True, text=True, timeout=90
    )
    base = os.path.splitext(os.path.basename(xlsx_path))[0]
    generated = os.path.join(temp_dir, f"{base}.pdf")
    if os.path.exists(generated):
        shutil.move(generated, out_pdf)
        return True
    print(f"  [WARN] Excel→PDF failed: {r.stderr[:200]}")
    return False

def merge_pdfs(pdf_list, output_path):
    writer = PdfWriter()
    for p in pdf_list:
        if os.path.exists(p):
            for page in PdfReader(p).pages:
                writer.add_page(page)
        else:
            print(f"  [WARN] Missing PDF: {p}")
    with open(output_path, "wb") as f:
        writer.write(f)

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="HLF Monthly Billing Package Builder")
    parser.add_argument("--billing-excel", required=True, help="Path to billing data .xlsx")
    parser.add_argument("--gst-bills-dir", required=True, help="Folder containing state GST bill PDFs")
    parser.add_argument("--output-dir", required=True, help="Where to save the final packages")
    parser.add_argument("--bill-date", required=True, help="Date for covering letter: DD.MM.YYYY")
    parser.add_argument("--billing-month", required=True, help="Month abbreviation: MAR, APR, etc.")
    parser.add_argument("--billing-year", required=True, type=int, help="Year: 2026")
    parser.add_argument("--states-only", default="", help="Comma-separated state names to process (optional)")
    parser.add_argument("--skip-missing-gst", action="store_true",
                        help="Skip states whose GST bill PDF is not found (default: auto-generate)")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    temp_dir = os.path.join(args.output_dir, "_temp")
    os.makedirs(temp_dir, exist_ok=True)

    # Load billing data
    df = pd.read_excel(args.billing_excel, sheet_name="Billing Data")
    df["Fees Amount"] = pd.to_numeric(df["Fees Amount"], errors="coerce").fillna(0)

    # Determine which states to process
    billing_states = sorted(df["State"].dropna().unique())
    if args.states_only:
        billing_states = [s.strip() for s in args.states_only.split(",")]

    print(f"\nHLF Monthly Billing Builder")
    print(f"Month: {args.billing_month} {args.billing_year}  |  Date: {args.bill_date}")
    print(f"States: {len(billing_states)}")
    print(f"Output: {args.output_dir}\n")

    # Check for new states not in master
    unknown = [s for s in billing_states if s not in STATES]
    if unknown:
        print(f"WARNING: These states are in the billing data but NOT in state_master.json:")
        for u in unknown:
            print(f"  - {u}")
        print("Add them to scripts/state_master.json before proceeding.\n")
        billing_states = [s for s in billing_states if s in STATES]

    results = []
    for state in billing_states:
        safe = state.replace(" ", "_")
        sd = STATES[state]
        print(f"  [{state}]")

        # ── Covering letter PDF
        cl_html = build_covering_letter_html(state, df, args.bill_date, args.billing_month, args.billing_year)
        cl_pdf = os.path.join(temp_dir, f"cl_{safe}.pdf")
        HTML(string=cl_html).write_pdf(cl_pdf)

        # ── GST bill PDF
        gst_pattern = sd["gst_file_pattern"].upper()
        gst_pdf = None
        # Try exact match, then case-insensitive
        for fname in os.listdir(args.gst_bills_dir):
            if fname.upper().replace(".PDF", "") == gst_pattern:
                gst_pdf = os.path.join(args.gst_bills_dir, fname)
                break

        if not gst_pdf:
            if args.skip_missing_gst:
                print(f"    [SKIP] No GST bill found for {state} — skipping")
                continue
            else:
                print(f"    [AUTO] No GST bill found for {state} — generating from billing data")
                gst_html = build_gst_bill_html(state, df, args.bill_date, args.billing_month, args.billing_year)
                gst_pdf = os.path.join(temp_dir, f"gst_{safe}.pdf")
                HTML(string=gst_html).write_pdf(gst_pdf)
                # Save a copy to GST bills folder for reference
                ref_path = os.path.join(args.gst_bills_dir, f"{gst_pattern}.pdf")
                shutil.copy(gst_pdf, ref_path)
                print(f"    [AUTO] Saved generated GST bill to: {ref_path}")
        else:
            print(f"    GST bill: {os.path.basename(gst_pdf)}")

        # ── Excel case detail PDF
        xl_path = build_excel_case_detail(state, df, temp_dir, args.billing_month, args.billing_year)
        xl_pdf = os.path.join(temp_dir, f"xl_{safe}.pdf")
        excel_to_pdf(xl_path, xl_pdf, temp_dir)

        # ── Merge
        out_path = os.path.join(args.output_dir, f"{safe}_Invoice_Package.pdf")
        merge_pdfs([cl_pdf, gst_pdf, xl_pdf], out_path)
        pages = len(PdfReader(out_path).pages)
        size_kb = os.path.getsize(out_path) // 1024
        print(f"    ✓ {safe}_Invoice_Package.pdf  ({pages} pages, {size_kb} KB)")
        results.append({"state": state, "file": f"{safe}_Invoice_Package.pdf", "pages": pages})

    # ── Cleanup temp
    shutil.rmtree(temp_dir, ignore_errors=True)

    # ── Summary
    print(f"\n{'─'*55}")
    print(f"Done! {len(results)} packages saved to:")
    print(f"  {args.output_dir}")
    print(f"{'─'*55}\n")
    return results

if __name__ == "__main__":
    main()
