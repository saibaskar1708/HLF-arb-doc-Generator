"""
generator.py
────────────
Core logic for reading the Excel file and rendering reference letters.
Imported by app.py (web) or run directly via generate_from_excel.py (CLI).
"""

import openpyxl
from docxtpl import DocxTemplate
from docx import Document
from docx.oxml import OxmlElement
from docx.oxml.ns import qn
from docx.shared import Inches
from datetime import datetime, date
from io import BytesIO
import copy
import os
import re
import zipfile


# ─── Helpers ──────────────────────────────────────────────────────────────────

def clean(val):
    """Safe string conversion — None and 'nan' become empty string."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s.lower() == "nan" else s


def clean_address(val):
    """Strip each line, drop blank lines, re-join with newline."""
    if not val:
        return ""
    lines = [ln.strip() for ln in str(val).split("\n") if ln.strip()]
    return "\n".join(lines)


def format_date(val):
    """Parse any date / datetime / string and return dd.MM.yyyy."""
    if not val:
        return ""
    if isinstance(val, (datetime, date)):
        return val.strftime("%d.%m.%Y")
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%d.%m.%Y")
        except ValueError:
            continue
    return s


def build_reg_chassis_phrase(reg_no, engine_no, chassis_no):
    """Combine vehicle identifiers into the template phrase."""
    parts = []
    if clean(reg_no):
        parts.append(f"Reg. No. {clean(reg_no)}")
    if clean(engine_no):
        parts.append(f"Engine No. {clean(engine_no)}")
    if clean(chassis_no):
        parts.append(f"Chassis No. {clean(chassis_no)}")
    return " | ".join(parts) if parts else "NA"


def safe_filename(s):
    """Replace characters invalid in filenames."""
    return re.sub(r'[\\/:*?"<>|]', "_", clean(s)).strip("_. ")


# ─── Numbered-list helpers ────────────────────────────────────────────────────

def _max_num_id(numbering_elem):
    ids = [
        int(n.get(qn("w:numId"), 0))
        for n in numbering_elem.findall(qn("w:num"))
    ]
    return max(ids, default=0)


def _remap_numbering(master_doc, sub_doc):
    try:
        master_np = master_doc.part.numbering_part._element
    except AttributeError:
        return {}
    try:
        sub_np = sub_doc.part.numbering_part._element
    except AttributeError:
        return {}

    next_id = _max_num_id(master_np) + 1
    id_map  = {}

    for sub_num in sub_np.findall(qn("w:num")):
        old_id  = sub_num.get(qn("w:numId"))
        new_num = copy.deepcopy(sub_num)
        new_num.set(qn("w:numId"), str(next_id))

        # Remove any existing lvlOverride elements, then add fresh ones that
        # force each numbering list to restart from 1 (prevents continuation
        # of numbering across letters in the combined document).
        for existing in new_num.findall(qn("w:lvlOverride")):
            new_num.remove(existing)
        for ilvl in range(9):
            lvl_override = OxmlElement("w:lvlOverride")
            lvl_override.set(qn("w:ilvl"), str(ilvl))
            start_override = OxmlElement("w:startOverride")
            start_override.set(qn("w:val"), "1")
            lvl_override.append(start_override)
            new_num.append(lvl_override)

        master_np.append(new_num)
        id_map[old_id] = str(next_id)
        next_id += 1

    return id_map


def _apply_num_id_remap(elem, id_map):
    for num_id_elem in elem.iter(qn("w:numId")):
        old_val = num_id_elem.get(qn("w:val"))
        if old_val in id_map:
            num_id_elem.set(qn("w:val"), id_map[old_val])


# ─── Rendering ────────────────────────────────────────────────────────────────

def render_to_buffer(template_path, context):
    """Render the Jinja2 docx template; return an in-memory BytesIO."""
    doc = DocxTemplate(template_path)
    doc.render(context)
    for section in doc.sections:
        section.left_margin   = Inches(0.5)
        section.right_margin  = Inches(0.5)
        section.top_margin    = Inches(0.5)
        section.bottom_margin = Inches(0.5)
    buf = BytesIO()
    doc.save(buf)
    buf.seek(0)
    return buf


# ─── Combiner ─────────────────────────────────────────────────────────────────

def combine_documents(buffers):
    """Merge a list of docx BytesIO buffers into a single Document."""
    if not buffers:
        return None

    buffers[0].seek(0)
    master = Document(buffers[0])

    for buf in buffers[1:]:
        buf.seek(0)
        sub = Document(buf)

        id_map = _remap_numbering(master, sub)

        master_body  = master.element.body
        final_sectPr = master_body.find(qn("w:sectPr"))

        def insert_before_sectPr(elem):
            if final_sectPr is not None:
                final_sectPr.addprevious(elem)
            else:
                master_body.append(elem)

        # Page-break between letters
        pg_p  = OxmlElement("w:p")
        pg_r  = OxmlElement("w:r")
        pg_br = OxmlElement("w:br")
        pg_br.set(qn("w:type"), "page")
        pg_r.append(pg_br)
        pg_p.append(pg_r)
        insert_before_sectPr(pg_p)

        for elem in sub.element.body:
            if elem.tag == qn("w:sectPr"):
                continue
            elem_copy = copy.deepcopy(elem)
            if id_map:
                _apply_num_id_remap(elem_copy, id_map)
            insert_before_sectPr(elem_copy)

    return master


# ─── Main generate function ───────────────────────────────────────────────────

def generate(
    excel_path: str,
    template_path: str,
    sheet_name: str    = "DataFormatted",
    ref_counter_start: int = 1,
    lot_label: str     = "Lot",
    progress_cb=None,  # optional callable(current, total, message)
) -> dict:
    """
    Read the Excel file, render one letter per row, and return:
    {
        "buffers":   [(filename, BytesIO), ...],    # individual letters
        "combined":  BytesIO,                       # all letters merged
        "zip":       BytesIO,                       # ZIP of individual files
        "success":   int,
        "skipped":   int,
        "errors":    [(row, contract_no, message), ...]
    }
    """
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
        )

    ws      = wb[sheet_name]
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    rows    = list(ws.iter_rows(min_row=2, values_only=True))
    total   = len(rows)

    buffers     = []        # (filename, BytesIO)
    ref_counter = ref_counter_start
    success     = 0
    skipped     = 0
    errors      = []

    for i, row in enumerate(rows, start=2):
        data = dict(zip(headers, row))

        contract_no = clean(data.get("Contract No"))
        if not contract_no:
            skipped += 1
            continue

        has_co_borrower = clean(data.get("HasCoBorrower", "")).upper() == "Y"
        has_guarantor   = clean(data.get("HasGuarantor",  "")).upper() == "Y"

        current_ref_no = f"HLF/SNS/REF/2026/{ref_counter}"

        context = {
            "current_ref_no":      current_ref_no,
            "date":                format_date(data.get("Reference Letter Date")),
            "contract_no":         contract_no,
            "agreement_date":      format_date(data.get("Contract Date")),
            "asset_description":   clean(data.get("Product Model") or data.get("Vehicle Details")),
            "reg_chassis_phrase":  build_reg_chassis_phrase(
                                       data.get("Vehicle No"),
                                       data.get("Engine No"),
                                       data.get("Chassis No"),
                                   ),
            "borrower_name":       clean(data.get("Borrower Name")),
            "borrower_address":    clean_address(data.get("Borrower Address")),
            "co_borrower_name":    clean(data.get("Co-Borrower Name"))              if has_co_borrower else "",
            "co_borrower_address": clean_address(data.get("Co-Borrower Address"))   if has_co_borrower else "",
            "guarantor_name":      clean(data.get("Guarantor Name"))                if has_guarantor else "",
            "guarantor_address":   clean_address(data.get("Guaranator Address"))    if has_guarantor else "",
            "lrn_ref_no":          clean(data.get("LRN Ref No")),
            "lrn_date":            format_date(data.get("LRN Date")),
            "claim_date":          format_date(data.get("Claim Date")),
            "claim_amount":        clean(data.get("Claim Amount")),
            "claim_amount_words":  clean(data.get("Claim Amount in Words")),
        }

        try:
            buf = render_to_buffer(template_path, context)
            fname = f"{safe_filename(current_ref_no)}_{safe_filename(contract_no)}.docx"
            buffers.append((fname, buf))
            ref_counter += 1
            success += 1
            if progress_cb:
                progress_cb(success + skipped, total, f"Generated: {current_ref_no}")
        except Exception as e:
            errors.append((i, contract_no, str(e)))

    # Build combined doc
    raw_bufs = [b for _, b in buffers]
    combined = combine_documents(raw_bufs)
    combined_buf = BytesIO()
    if combined:
        combined.save(combined_buf)
        combined_buf.seek(0)

    # Build ZIP of individual files
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, buf in buffers:
            buf.seek(0)
            zf.writestr(fname, buf.read())
        # Also include the combined file inside the ZIP
        if combined:
            combined_buf.seek(0)
            zf.writestr(f"Combined_Reference_Letters_{lot_label}.docx", combined_buf.read())
    zip_buf.seek(0)
    combined_buf.seek(0)

    return {
        "buffers":  buffers,
        "combined": combined_buf,
        "zip":      zip_buf,
        "success":  success,
        "skipped":  skipped,
        "errors":   errors,
        "total":    total,
    }
