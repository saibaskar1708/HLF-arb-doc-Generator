"""
app.py  –  HLF Reference Letter Generator  (Flask web app)
───────────────────────────────────────────────────────────
Run:
    python app.py

Then open  http://localhost:5000  in your browser.
"""

import os
import json
import uuid
import zipfile
import threading
from io import BytesIO
from datetime import datetime

import openpyxl
from flask import (
    Flask, request, render_template, redirect, url_for,
    send_file, jsonify, flash, session
)
from werkzeug.utils import secure_filename

from generator import generate
from generate_noh import (
    build_noh, merge_noh_docs,
    C_CONTRACT_NO, C_BORROWER_NAME, C_NOH_DATE, C_MEETING_LINK, C_ARB_NAME, C_CASE_NO,
    v as noh_v,
)

# ─── App setup ────────────────────────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.urandom(24)

UPLOAD_FOLDER   = "uploads"
OUTPUT_FOLDER   = "outputs"
NOH_FOLDER      = "NOH_Output"
TEMPLATE_PATH   = os.path.join("letter_templates", "Reference_Letter_Template.docx")
ALLOWED_EXT     = {".xlsx", ".xls"}

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(NOH_FOLDER,    exist_ok=True)

# In-memory job store  {job_id: {...}}
jobs: dict[str, dict] = {}


# ─── Helpers ──────────────────────────────────────────────────────────────────

def allowed_file(filename):
    return os.path.splitext(filename)[1].lower() in ALLOWED_EXT


def run_job(job_id, excel_path, sheet_name, ref_start, lot_label):
    """Background thread: runs generation and stores result in jobs dict."""
    job = jobs[job_id]
    job["status"] = "running"
    job["progress"] = 0

    def progress_cb(current, total, msg):
        job["progress"] = round(current / total * 100) if total else 100
        job["message"]  = msg

    try:
        result = generate(
            excel_path    = excel_path,
            template_path = TEMPLATE_PATH,
            sheet_name    = sheet_name,
            ref_counter_start = ref_start,
            lot_label     = lot_label,
            progress_cb   = progress_cb,
        )

        # Persist ZIP to disk so we can serve it later
        zip_filename = f"HLF_Reference_Letters_{lot_label}_{job_id[:8]}.zip"
        zip_path = os.path.join(OUTPUT_FOLDER, zip_filename)
        with open(zip_path, "wb") as f:
            result["zip"].seek(0)
            f.write(result["zip"].read())

        # Persist combined DOCX
        combined_filename = f"Combined_Reference_Letters_{lot_label}_{job_id[:8]}.docx"
        combined_path = os.path.join(OUTPUT_FOLDER, combined_filename)
        with open(combined_path, "wb") as f:
            result["combined"].seek(0)
            f.write(result["combined"].read())

        job.update({
            "status":           "done",
            "progress":         100,
            "success":          result["success"],
            "skipped":          result["skipped"],
            "errors":           result["errors"],
            "total":            result["total"],
            "zip_filename":     zip_filename,
            "combined_filename": combined_filename,
            "message":          f"Done! {result['success']} letters generated.",
        })

    except Exception as e:
        job.update({
            "status":  "error",
            "message": str(e),
        })
    finally:
        # Clean up upload
        try:
            os.remove(excel_path)
        except OSError:
            pass


def run_noh_job(job_id, excel_path, sheet_name, sig_map=None):
    """Background thread: generates NOH documents and stores result in jobs dict.

    sig_map: dict of {arbitrator_name_lower: temp_image_path}
    """
    job = jobs[job_id]
    job["status"]   = "running"
    job["progress"] = 0

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        all_rows  = list(ws.iter_rows(values_only=True))
        wb.close()

        data_rows = [r for r in all_rows[1:] if any(c is not None for c in r)]
        total     = len(data_rows)

        generated     = 0
        skipped_list  = []
        errors        = []
        doc_bufs      = []   # collect for combined DOCX

        zip_buf = BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for i, row_data in enumerate(data_rows):
                contract_no   = noh_v(row_data[C_CONTRACT_NO])
                borrower_name = noh_v(row_data[C_BORROWER_NAME])
                case_no       = noh_v(row_data[C_CASE_NO])

                job["progress"] = round((i + 1) / total * 100) if total else 100
                job["message"]  = f"Processing {i + 1}/{total} — {contract_no}"

                # Mandatory field check
                missing = []
                if not row_data[C_NOH_DATE]:
                    missing.append("NOH Date")
                if not noh_v(row_data[C_MEETING_LINK]):
                    missing.append("Meeting Link")

                if missing:
                    skipped_list.append({
                        "contract_no":   contract_no,
                        "borrower_name": borrower_name,
                        "reason":        f"Missing: {', '.join(missing)}",
                    })
                    continue

                try:
                    # Look up signature by arbitrator name (case-insensitive)
                    arb_name_key = noh_v(row_data[C_ARB_NAME]).lower().strip()
                    sig_path = (sig_map or {}).get(arb_name_key)
                    doc = build_noh(row_data, sig_image_path=sig_path)

                    # Filename: case_no first (for sorting), then contract_no
                    safe_case     = case_no.replace("/", "_").replace("\\", "_")
                    safe_contract = contract_no.replace("/", "_").replace("\\", "_")
                    fname   = f"NOH_{safe_case}_{safe_contract}.docx"

                    doc_buf = BytesIO()
                    doc.save(doc_buf)
                    doc_buf.seek(0)
                    doc_bytes = doc_buf.read()
                    zf.writestr(fname, doc_bytes)
                    doc_bufs.append(BytesIO(doc_bytes))   # keep for combined
                    generated += 1
                except Exception as e:
                    errors.append([i + 2, contract_no, str(e)])

        # Persist ZIP
        zip_filename = f"HLF_NOH_{job_id[:8]}.zip"
        zip_path     = os.path.join(NOH_FOLDER, zip_filename)
        zip_buf.seek(0)
        with open(zip_path, "wb") as f:
            f.write(zip_buf.read())

        # Build and persist combined DOCX
        combined_filename = None
        if doc_bufs:
            combined_doc = merge_noh_docs(doc_bufs)
            combined_filename = f"HLF_NOH_Combined_{job_id[:8]}.docx"
            combined_path = os.path.join(NOH_FOLDER, combined_filename)
            combined_doc.save(combined_path)

        job.update({
            "status":            "done",
            "progress":          100,
            "doc_type":          "noh",
            "success":           generated,
            "skipped":           len(skipped_list),
            "skipped_details":   skipped_list,
            "errors":            errors,
            "total":             total,
            "zip_filename":      zip_filename,
            "combined_filename": combined_filename,
            "message":           f"Done! {generated} NOH document(s) generated.",
        })

    except Exception as e:
        job.update({"status": "error", "message": str(e)})
    finally:
        try:
            os.remove(excel_path)
        except OSError:
            pass
        for sig_path in (sig_map or {}).values():
            try:
                os.remove(sig_path)
            except OSError:
                pass


# ─── Routes ───────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/scan-arbitrators", methods=["POST"])
def scan_arbitrators():
    """Read uploaded Excel and return unique arbitrator names from C_ARB_NAME column."""
    f = request.files.get("excel_file")
    if not f or not f.filename:
        return jsonify({"arbitrators": []})
    sheet_name = request.form.get("sheet_name", "DataFormatted").strip() or "DataFormatted"
    try:
        buf = BytesIO(f.read())
        wb  = openpyxl.load_workbook(buf, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return jsonify({"arbitrators": [], "error": f"Sheet '{sheet_name}' not found"})
        ws   = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        names = sorted(set(
            str(r[C_ARB_NAME]).strip()
            for r in rows[1:] if r[C_ARB_NAME] is not None and str(r[C_ARB_NAME]).strip()
        ))
        return jsonify({"arbitrators": names})
    except Exception as e:
        return jsonify({"arbitrators": [], "error": str(e)})


@app.route("/generate", methods=["POST"])
def start_generate():
    """Receives the form, saves the Excel, starts background job."""
    if "excel_file" not in request.files:
        flash("No file uploaded.", "danger")
        return redirect(url_for("index"))

    f = request.files["excel_file"]
    if not f.filename or not allowed_file(f.filename):
        flash("Please upload a valid .xlsx or .xls file.", "danger")
        return redirect(url_for("index"))

    doc_type   = request.form.get("doc_type", "reference_letter")
    sheet_name = request.form.get("sheet_name", "DataFormatted").strip() or "DataFormatted"

    # Save uploaded file
    fname      = secure_filename(f.filename)
    excel_path = os.path.join(UPLOAD_FOLDER, f"{uuid.uuid4().hex}_{fname}")
    f.save(excel_path)

    # Create job
    job_id = uuid.uuid4().hex
    jobs[job_id] = {
        "status":   "queued",
        "progress": 0,
        "message":  "Starting…",
        "doc_type": doc_type,
        "created":  datetime.now().strftime("%H:%M:%S"),
    }

    if doc_type == "noh":
        # Collect arbitrator name → signature file mappings (sig_name_0/sig_image_0, …)
        sig_map = {}
        for idx in range(20):
            sig_name = request.form.get(f"sig_name_{idx}", "").strip()
            sig_file = request.files.get(f"sig_image_{idx}")
            if not sig_name and not (sig_file and sig_file.filename):
                continue
            if sig_name and sig_file and sig_file.filename:
                sig_ext = os.path.splitext(secure_filename(sig_file.filename))[1].lower()
                if sig_ext in {".png", ".jpg", ".jpeg"}:
                    sig_path = os.path.join(UPLOAD_FOLDER, f"{uuid.uuid4().hex}_sig{sig_ext}")
                    sig_file.save(sig_path)
                    sig_map[sig_name.lower()] = sig_path
        t = threading.Thread(
            target=run_noh_job,
            args=(job_id, excel_path, sheet_name),
            kwargs={"sig_map": sig_map},
            daemon=True,
        )
    else:
        lot_label = request.form.get("lot_label", "Lot").strip() or "Lot"
        try:
            ref_start = int(request.form.get("ref_start", "1"))
        except ValueError:
            ref_start = 1
        jobs[job_id]["lot_label"] = lot_label
        t = threading.Thread(
            target=run_job,
            args=(job_id, excel_path, sheet_name, ref_start, lot_label),
            daemon=True,
        )

    t.start()
    return redirect(url_for("status_page", job_id=job_id))


@app.route("/status/<job_id>")
def status_page(job_id):
    job = jobs.get(job_id)
    if not job:
        flash("Job not found.", "danger")
        return redirect(url_for("index"))
    return render_template("status.html", job_id=job_id, job=job)


@app.route("/api/status/<job_id>")
def api_status(job_id):
    job = jobs.get(job_id)
    if not job:
        return jsonify({"status": "not_found"}), 404
    return jsonify(job)


@app.route("/download/<job_id>/<filetype>")
def download(job_id, filetype):
    job = jobs.get(job_id)
    if not job or job.get("status") != "done":
        flash("File not ready or job not found.", "danger")
        return redirect(url_for("index"))

    doc_type = job.get("doc_type", "reference_letter")

    if filetype == "zip":
        folder = NOH_FOLDER if doc_type == "noh" else OUTPUT_FOLDER
        path   = os.path.join(folder, job["zip_filename"])
        return send_file(path, as_attachment=True,
                         download_name=job["zip_filename"],
                         mimetype="application/zip")
    elif filetype == "combined":
        if not job.get("combined_filename"):
            flash("Combined file not available.", "danger")
            return redirect(url_for("index"))
        folder = NOH_FOLDER if doc_type == "noh" else OUTPUT_FOLDER
        path   = os.path.join(folder, job["combined_filename"])
        return send_file(path, as_attachment=True,
                         download_name=job["combined_filename"],
                         mimetype="application/vnd.openxmlformats-officedocument.wordprocessingml.document")
    else:
        flash("Unknown file type.", "danger")
        return redirect(url_for("index"))


# ─── Run ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("─" * 55)
    print("  HLF Reference Letter Generator")
    print("  Open http://localhost:5000 in your browser")
    print("─" * 55)
    app.run(debug=True, port=5000)
